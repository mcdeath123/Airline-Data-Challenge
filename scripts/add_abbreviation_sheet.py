"""
add_abbreviation_sheet.py

This script adds a new, beautifully formatted sheet called 'Abbreviation Definitions'
to the Airline_Analysis_Results.xlsx Excel workbook. It provides definitions
for all airport terminal IATA codes and engineered metadata column headers.
Handles locked file scenarios gracefully by writing to a fallback filename.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_abbreviation_sheet():
    excel_file = "Airline_Analysis_Results.xlsx"
    fallback_file = "Airline_Analysis_Results_New.xlsx"
    
    print("Loading or creating Excel workbook...")
    if os.path.exists(excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file)
            print(f"Loaded existing workbook: {excel_file}")
        except Exception as e:
            print(f"Error loading {excel_file}: {e}. Creating new workbook.")
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet if brand new
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
    sheet_name = "Abbreviation Definitions"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        print(f"Removed existing sheet '{sheet_name}' to overwrite.")
        
    # Create the sheet at position 0 (first tab)
    ws = wb.create_sheet(title=sheet_name, index=0)
    ws.views.sheetView[0].showGridLines = True
    
    # Define color palette (matching the other sheets - elegant deep navy)
    primary_color = "1E3D59"   # Sleek deep navy
    accent_color = "FF6E40"    # Sunrise orange accent
    light_fill_color = "F4F6F9" # Light silver-grey for data rows
    accent_fill_color = "FFF5F2" # Very soft orange highlight
    
    # Styles
    title_font = Font(name="Trebuchet MS", size=16, bold=True, color="1E3D59")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="555555")
    section_font = Font(name="Trebuchet MS", size=13, bold=True, color="1E3D59")
    header_font = Font(name="Trebuchet MS", size=10.5, bold=True, color="FFFFFF")
    data_font_bold = Font(name="Segoe UI", size=10, bold=True, color="333333")
    data_font_regular = Font(name="Segoe UI", size=10, color="333333")
    
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    zebra_fill = PatternFill(start_color=light_fill_color, end_color=light_fill_color, fill_type="solid")
    accent_fill = PatternFill(start_color=accent_fill_color, end_color=accent_fill_color, fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thick_bottom_side = Side(border_style="medium", color="1E3D59")
    
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # 1. Main Sheet Title
    ws["B2"] = "Airline Data Challenge: Metadata Definitions"
    ws["B2"].font = title_font
    ws["B3"] = "Reference definitions for airport terminal abbreviations (IATA codes) and engineered data column headers used in the Q1 2019 analysis."
    ws["B3"].font = subtitle_font
    
    # 2. Section 1: Metadata Column Headers Definitions
    ws["B5"] = "I. Analytical Data Schema & Column Headers"
    ws["B5"].font = section_font
    
    headers_metadata = ["Column Header", "Data Type", "Definition / Formula", "Data Source / Reference"]
    col_idx_meta = 2 # Start from column B
    row_idx_meta = 6
    
    # Write metadata headers
    for idx, h in enumerate(headers_metadata):
        cell = ws.cell(row=row_idx_meta, column=col_idx_meta + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border
        
    ws.row_dimensions[row_idx_meta].height = 26
    
    # Metadata data definitions
    metadata_rows = [
        ["route_undir", "String", "Undirected route key combining two alphabetically sorted 3-letter IATA airport codes (e.g. 'JFK-LAX').", "Engineered Key"],
        ["flights_count", "Integer", "Total count of individual, non-cancelled flight legs operated in Q1 2019.", "Flights.csv"],
        ["round_trips_count", "Float", "Total number of round-trip flights operated in Q1 2019 (flights_count / 2).", "Calculated Metric"],
        ["total_passengers", "Float", "Sum of all passengers carried across all legs (capacity of 200 * Occupancy Rate).", "Calculated Metric"],
        ["average_occupancy", "Float", "Mean occupancy rate (%) of flights operated on the route.", "Flights.csv"],
        ["average_distance", "Float", "Mean physical flight distance in miles.", "Flights.csv"],
        ["average_round_trip_fare", "Float", "Mean round-trip ticket fare ($) from the Tickets dataset, excluding reward security fees (<$50) and extreme anomalies (>$5000).", "Tickets.csv"],
        ["total_ticket_revenue", "Float", "Total ticket sales revenue: passengers * (average_round_trip_fare / 2) per flight leg.", "Calculated Metric"],
        ["total_baggage_revenue", "Float", "Total ancillary checked bag revenue: passengers * 50% check rate * $35.00 leg bag fee.", "Calculated Metric"],
        ["total_revenue", "Float", "Combined total revenue: total_ticket_revenue + total_baggage_revenue.", "Calculated Metric"],
        ["total_distance_cost", "Float", "Physical operating mileage cost: distance * $9.18 per mile ($8.00 fuel/crew/maintenance + $1.18 depreciation/insurance).", "Calculated Metric"],
        ["total_airport_cost", "Float", "Fixed landing operational fees: $10,000 for Large, $5,000 for Medium destinations.", "Calculated Metric"],
        ["total_delay_cost", "Float", "Punctuality penalty: $75.00 per minute for arrival and departure delays over 15 minutes.", "Calculated Metric"],
        ["total_cost", "Float", "Combined total operational cost: total_distance_cost + total_airport_cost + total_delay_cost.", "Calculated Metric"],
        ["total_profit", "Float", "Combined net operating profit: total_revenue - total_cost.", "Calculated Metric"],
        ["profit_per_round_trip", "Float", "Net profit generated per round-trip flight: total_profit / round_trips_count.", "Calculated Metric"],
        ["revenue_per_round_trip", "Float", "Total revenue generated per round-trip flight: total_revenue / round_trips_count.", "Calculated Metric"],
        ["cost_per_round_trip", "Float", "Total operational cost per round-trip flight: total_cost / round_trips_count.", "Calculated Metric"],
        ["breakeven_round_trips", "Float", "Minimum round-trip flights needed to cover $90,000,000 upfront aircraft CapEx ($90M / profit_per_round_trip).", "Calculated Metric"]
    ]
    
    current_row = row_idx_meta + 1
    for r_idx, row_data in enumerate(metadata_rows):
        ws.row_dimensions[current_row].height = 20
        # Determine filling for zebra stripes
        row_fill = zebra_fill if r_idx % 2 == 1 else None
        
        for c_offset, val in enumerate(row_data):
            cell = ws.cell(row=current_row, column=col_idx_meta + c_offset)
            cell.value = val
            cell.border = data_border
            if row_fill:
                cell.fill = row_fill
            
            # Formatting specifics
            if c_offset == 0:
                cell.font = data_font_bold
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = data_font_regular
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
        current_row += 1
        
    # 3. Section 2: Terminal IATA Codes
    current_row += 2 # Add spacing
    ws.cell(row=current_row, column=2).value = "II. Airport Terminal Abbreviations (IATA Codes)"
    ws.cell(row=current_row, column=2).font = section_font
    
    current_row += 1
    headers_iata = ["IATA Code", "Airport Name / Terminal Description", "Served City", "State", "Airport Size Classification"]
    for idx, h in enumerate(headers_iata):
        cell = ws.cell(row=current_row, column=col_idx_meta + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = header_border
        
    ws.row_dimensions[current_row].height = 26
    
    iata_rows = [
        ["JFK", "John F. Kennedy International Airport", "New York", "NY", "Large Airport"],
        ["LAX", "Los Angeles International Airport", "Los Angeles", "CA", "Large Airport"],
        ["SFO", "San Francisco International Airport", "San Francisco", "CA", "Large Airport"],
        ["ORD", "Chicago O'Hare International Airport", "Chicago", "IL", "Large Airport"],
        ["LGA", "LaGuardia Airport", "New York", "NY", "Large Airport"],
        ["ATL", "Hartsfield-Jackson Atlanta International Airport", "Atlanta", "GA", "Large Airport"],
        ["CLT", "Charlotte Douglas International Airport", "Charlotte", "NC", "Large Airport"],
        ["EWR", "Newark Liberty International Airport", "Newark", "NJ", "Large Airport"],
        ["DCA", "Ronald Reagan Washington National Airport", "Washington", "DC", "Large Airport"],
        ["FLO", "Florence Regional Airport", "Florence", "SC", "Medium Airport"],
        ["SEA", "Seattle-Tacoma International Airport", "Seattle", "WA", "Large Airport"],
        ["OGG", "Kahului Airport", "Kahului", "HI", "Medium Airport"],
        ["HNL", "Daniel K. Inouye International Airport", "Honolulu", "HI", "Large Airport"],
        ["PDX", "Portland International Airport", "Portland", "OR", "Large Airport"],
        ["BOS", "Boston Logan International Airport", "Boston", "MA", "Large Airport"],
        ["MCO", "Orlando International Airport", "Orlando", "FL", "Large Airport"]
    ]
    
    current_row += 1
    for r_idx, row_data in enumerate(iata_rows):
        ws.row_dimensions[current_row].height = 20
        # Determine filling for zebra stripes, highlight medium airports slightly differently
        is_medium = row_data[4] == "Medium Airport"
        row_fill = accent_fill if is_medium else (zebra_fill if r_idx % 2 == 1 else None)
        
        for c_offset, val in enumerate(row_data):
            cell = ws.cell(row=current_row, column=col_idx_meta + c_offset)
            cell.value = val
            cell.border = data_border
            if row_fill:
                cell.fill = row_fill
                
            if c_offset == 0:
                cell.font = data_font_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = data_font_regular
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        current_row += 1
        
    # Auto-adjust column widths with generous margins to avoid any clipping
    column_widths = {
        "B": 24, # Column Header / IATA Code
        "C": 38, # Data Type / Airport Name
        "D": 85, # Definition / City
        "E": 24, # Source / State
        "F": 28  # Airport Size (Section II only)
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Save file with robust lock handling
    print(f"Saving workbook as {excel_file}...")
    try:
        wb.save(excel_file)
        print("Workbook successfully saved!")
    except PermissionError:
        print(f"WARNING: Permission denied when saving to {excel_file}. The file might be open in Excel. Saving to fallback: {fallback_file}...")
        wb.save(fallback_file)
        print(f"Workbook successfully saved to fallback: {fallback_file}!")

if __name__ == "__main__":
    create_abbreviation_sheet()
